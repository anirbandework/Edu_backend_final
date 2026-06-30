"""Bulk import of members from an Excel/CSV file, per role.

Flow: the admin picks a role → downloads a template whose columns are the built-in
identity fields + that role's custom fields → fills it (up to ~thousands of rows) →
uploads it. We parse (streaming for .xlsx), validate every row, skip duplicate phones,
and bulk-insert the good rows in chunks, returning a per-row result report.
"""
from __future__ import annotations

import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.member import Member
from ...auth_rbac.access import custom_fields as cf
from ...authority_management.models.authority import Authority

MAX_ROWS = 5000          # hard cap per upload
MAX_BYTES = 10 * 1024 * 1024  # 10 MB
_CHUNK = 500             # insert/commit batch size
_DV_MAX = 250            # max length of an inline dropdown list (Excel formula1 limit ~255)


def _gen_staff_id() -> str:
    return "STF-" + uuid.uuid4().hex[:12].upper()  # 48 bits — negligible collision in a bulk run


def import_columns(role) -> list[dict]:
    """Ordered column spec for a role's template: built-in identity fields, then the
    role's custom fields (key prefixed 'cf:')."""
    cols = [
        {"key": "first_name", "label": "First name", "required": True, "type": "text"},
        {"key": "last_name", "label": "Last name", "required": True, "type": "text"},
        {"key": "phone", "label": "Phone", "required": True, "type": "phone"},
        {"key": "email", "label": "Email", "required": False, "type": "email"},
        {"key": "position", "label": "Designation", "required": False, "type": "text"},
    ]
    for f in (role.custom_fields or []):
        cols.append({
            "key": "cf:" + f["key"], "label": f.get("label", f["key"]),
            "required": bool(f.get("required")), "type": f.get("type", "text"),
            "options": f.get("options"),
        })
    return cols


def _norm(h) -> str:
    return str(h or "").strip().rstrip("*").strip().lower()


def build_template_xlsx(role) -> bytes:
    """A .xlsx template: a 'Users' sheet (bold header, frozen, dropdowns for select
    fields) + an 'Instructions' sheet describing every column."""
    cols = import_columns(role)
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    for ci, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=c["label"] + (" *" if c["required"] else ""))
        cell.font = Font(bold=True)
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = max(14, len(c["label"]) + 5)
        # Force phone columns to TEXT so Excel preserves leading zeros (a phone is the
        # login id — losing a leading 0 would lock the user out). Column-level format
        # applies to empty cells without bloating the file with thousands of styled cells.
        if c["type"] == "phone":
            ws.column_dimensions[letter].number_format = "@"
        if c["type"] == "select" and c.get("options"):
            joined = ",".join(c["options"])
            if len(joined) <= _DV_MAX:  # inline list (long lists fall back to free text)
                dv = DataValidation(type="list", formula1=f'"{joined}"',
                                    allow_blank=not c["required"])
                ws.add_data_validation(dv)
                dv.add(f"{letter}2:{letter}{MAX_ROWS + 1}")
    ws.freeze_panes = "A2"
    # Stamp the role id so an upload under a DIFFERENT role is caught, not silently mis-mapped.
    wb.properties.category = f"eduassist-role:{role.id}"

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Required", "Type", "Allowed values / notes"])
    for c in info[1]:
        c.font = Font(bold=True)
    for c in cols:
        note = ", ".join(c.get("options") or [])
        if not note and c["key"] == "phone":
            note = "Login id — must be unique"
        info.append([c["label"], "Yes" if c["required"] else "No", c["type"], note])
    for col in ("A", "B", "C", "D"):
        info.column_dimensions[col].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(v) -> str:
    """Coerce a spreadsheet cell to a trimmed string (numbers → no trailing .0, so a
    phone typed as a number stays '9999999999')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def parse_rows(filename: str, content: bytes, role) -> tuple[list[dict], str | None]:
    """Parse an uploaded .xlsx/.csv into row dicts keyed by built-in field + custom_raw.
    Returns (rows, error). Maps file columns to fields by header label (so column order
    can vary); unrecognised columns are ignored."""
    cols = import_columns(role)
    # Built-in columns WIN over a custom field that happens to share a label (iterate so
    # built-ins are assigned last). normalize_definitions also blocks such labels upstream.
    by_label = {_norm(c["label"]): c for c in reversed(cols)}

    name = (filename or "").lower()
    table: list[list] = []
    if name.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        table = [row for row in csv.reader(io.StringIO(text))]
    else:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            return [], "Could not read the file. Upload the .xlsx template or a .csv."
        # Guard against uploading a DIFFERENT role's template (stamped in build_template_xlsx).
        stamp = getattr(wb.properties, "category", None) or ""
        if stamp.startswith("eduassist-role:") and stamp.split(":", 1)[1] != str(role.id):
            return [], ("This template was generated for a different role. Download the "
                        "template for the selected role and use that.")
        ws = wb.active
        table = [list(r) for r in ws.iter_rows(values_only=True)]

    # find the header row (first non-empty)
    header_idx = next((i for i, r in enumerate(table) if any(_cell_str(c) for c in r)), None)
    if header_idx is None:
        return [], "The file is empty."
    headers = table[header_idx]
    # map file column index -> field spec
    col_field: dict[int, dict] = {}
    for idx, h in enumerate(headers):
        spec = by_label.get(_norm(h))
        if spec:
            col_field[idx] = spec
    if not any(col_field.get(i, {}).get("key") in ("first_name", "last_name", "phone")
               for i in col_field):
        return [], ("Template headers not recognised. Download the template for this "
                    "role and fill it without renaming the column headers.")

    rows: list[dict] = []
    for r_i in range(header_idx + 1, len(table)):
        raw = table[r_i]
        if not any(_cell_str(c) for c in raw):
            continue  # skip blank lines
        rec = {"row_no": r_i + 1, "first_name": "", "last_name": "", "phone": "",
               "email": "", "position": "", "custom_raw": {}}
        for idx, spec in col_field.items():
            val = _cell_str(raw[idx]) if idx < len(raw) else ""
            key = spec["key"]
            if key.startswith("cf:"):
                rec["custom_raw"][key[3:]] = val
            else:
                rec[key] = val
        rows.append(rec)
    return rows, None


async def bulk_import(db: AsyncSession, *, organisation_id, role, rows, created_by) -> dict:
    """Validate, dedupe (phone + email, in-file + existing) and bulk-insert. Returns
    {created, skipped:[{row,reason}], failed:[{row,reason}], total}."""
    created = 0
    failed: list[dict] = []
    skipped: list[dict] = []
    valid: list[dict] = []
    seen_phones: set[str] = set()
    seen_emails: set[str] = set()

    for r in rows:
        rn = r["row_no"]
        fn, ln, ph = r["first_name"].strip(), r["last_name"].strip(), r["phone"].strip()
        em = (r["email"] or "").strip() or None
        pos = (r["position"] or "").strip() or None
        if not (fn and ln and ph):
            failed.append({"row": rn, "reason": "First name, last name and phone are required"})
            continue
        if em and not cf._EMAIL_RE.match(em):
            failed.append({"row": rn, "reason": f"Invalid email: {em}"})
            continue
        try:
            clean = cf.validate_values(role.custom_fields or [], r["custom_raw"])
        except ValueError as e:
            failed.append({"row": rn, "reason": str(e)})
            continue
        if ph in seen_phones:
            skipped.append({"row": rn, "reason": f"Duplicate phone in file: {ph}"})
            continue
        if em and em.lower() in seen_emails:
            failed.append({"row": rn, "reason": f"Duplicate email in file: {em}"})
            continue
        seen_phones.add(ph)
        if em:
            seen_emails.add(em.lower())
        valid.append({"row": rn, "fn": fn, "ln": ln, "ph": ph, "em": em, "pos": pos, "clean": clean})

    # existing phones (members + authorities) and emails (members), in one query each
    phones = [v["ph"] for v in valid]
    existing_ph: set[str] = set()
    if phones:
        m = await db.execute(select(Member.phone).where(
            Member.phone.in_(phones), Member.is_deleted == False))  # noqa: E712
        a = await db.execute(select(Authority.phone).where(
            Authority.phone.in_(phones), Authority.is_deleted == False))  # noqa: E712
        existing_ph = {p for (p,) in m.all()} | {p for (p,) in a.all()}
    emails = [v["em"] for v in valid if v["em"]]
    existing_em: set[str] = set()
    if emails:
        me = await db.execute(select(Member.email).where(
            Member.email.in_(emails), Member.is_deleted == False))  # noqa: E712
        existing_em = {(e or "").lower() for (e,) in me.all()}

    to_insert: list[dict] = []
    for v in valid:
        if v["ph"] in existing_ph:
            skipped.append({"row": v["row"], "reason": f"Phone already registered: {v['ph']}"})
        elif v["em"] and v["em"].lower() in existing_em:
            failed.append({"row": v["row"], "reason": f"Email already in use: {v['em']}"})
        else:
            to_insert.append(v)

    def _make(v) -> Member:
        return Member(
            organisation_id=organisation_id, rbac_role_id=role.id, staff_id=_gen_staff_id(),
            first_name=v["fn"], last_name=v["ln"], phone=v["ph"], email=v["em"],
            position=v["pos"], status="invited", role="staff", created_by=created_by,
            profile={"custom_fields": v["clean"]} if v["clean"] else None)

    for i in range(0, len(to_insert), _CHUNK):
        chunk = to_insert[i:i + _CHUNK]
        db.add_all([_make(v) for v in chunk])
        try:
            await db.commit()
            created += len(chunk)
        except IntegrityError:
            # Rare race (a phone/email registered between our check and commit) — retry
            # the chunk row-by-row so one bad row doesn't sink the rest.
            await db.rollback()
            for v in chunk:
                db.add(_make(v))
                try:
                    await db.commit()
                    created += 1
                except IntegrityError:
                    await db.rollback()
                    failed.append({"row": v["row"], "reason": "Phone or email already in use"})

    failed.sort(key=lambda x: x["row"])
    skipped.sort(key=lambda x: x["row"])
    return {"created": created, "skipped": skipped, "failed": failed, "total": len(rows)}
